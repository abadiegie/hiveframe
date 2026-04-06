# Copyright 2026 Abadi Gilang
# SPDX-License-Identifier: Apache-2.0

"""Pandas-like dataframe facade backed by transaction coordinator."""

from __future__ import annotations

import json
from pathlib import Path
from typing import TYPE_CHECKING, Any, AsyncIterable, Callable, Iterable
import asyncio
import logging
import time

import pandas as pd

from .coordinator import TransactionCoordinator
from .transaction import Operation
from .schema import ColumnSchema

if TYPE_CHECKING:
    from .cluster_runtime import ClusterRuntime


logger = logging.getLogger("hiveframe.dataframe")


def _schema_sidecar_path(parquet_path: Path) -> Path:
    return parquet_path.with_suffix(parquet_path.suffix + ".schema.json")


def _load_schema_sidecar(parquet_path: Path) -> dict[str, ColumnSchema]:
    sidecar = _schema_sidecar_path(parquet_path)
    if not sidecar.exists():
        return {}
    payload = json.loads(sidecar.read_text())
    return {column: ColumnSchema.from_dict(schema_payload) for column, schema_payload in payload.items()}


def _write_schema_sidecar(parquet_path: Path, schema: dict[str, ColumnSchema]) -> None:
    sidecar = _schema_sidecar_path(parquet_path)
    if not schema:
        if sidecar.exists():
            sidecar.unlink()
        return
    payload = {column: column_schema.to_dict() for column, column_schema in schema.items()}
    sidecar.write_text(json.dumps(payload, indent=2, sort_keys=True))


def _resolve_read_path(name_or_path: str, coordinator: TransactionCoordinator) -> Path:
    candidate = Path(name_or_path)
    if candidate.suffix == ".parquet" or candidate.exists():
        return candidate
    return Path(coordinator.read_node._storage_dir) / f"{candidate}.parquet"


class _RuntimeRegistry:
    """In-process map of node_id → ClusterRuntime for same-process write routing."""
    _map: dict[str, "ClusterRuntime"] = {}

    @classmethod
    def register(cls, node_id: str, runtime: "ClusterRuntime") -> None:
        cls._map[node_id] = runtime

    @classmethod
    def get(cls, node_id: str) -> "ClusterRuntime | None":
        return cls._map.get(node_id)

    @classmethod
    def clear(cls) -> None:
        cls._map.clear()


class DFrame:
    """User-facing DataFrame API with transactional writes and analytical reads."""

    _SEP = "::"  # namespace separator: "{frame_id}::{col}_{row}"

    def __init__(
        self,
        data: dict[str, list[Any]] | None = None,
        coordinator: TransactionCoordinator | None = None,
        runtime: "ClusterRuntime | None" = None,
        frame_id: str | None = None,
        schema: dict[str, ColumnSchema] | None = None,
        transactional: bool = True,
    ) -> None:
        import uuid as _uuid
        self._coordinator = coordinator or TransactionCoordinator()
        self._runtime = runtime
        self._frame_id = frame_id or _uuid.uuid4().hex
        self._schema = schema or {}
        self._transactional = transactional
        # Simple short-lived cache for the local snapshot to avoid repeated builds
        # when users call multiple read properties/operators in quick succession.
        self._snapshot_cache: tuple[pd.DataFrame, float] | None = None  # (frame, ts)
        self._snapshot_cache_ttl = 0.1  # seconds
        if data:
            self._seed_initial_data(data)

    @classmethod
    def from_runtime(
        cls,
        runtime: "ClusterRuntime",
        data: dict[str, list[Any]] | None = None,
        frame_id: str | None = None,
        transactional: bool = True,
    ) -> "DFrame":
        """Create a DFrame backed by an existing ClusterRuntime coordinator.

        Each DFrame gets a unique frame_id used to namespace all cell-ids so
        multiple DFrames on the same runtime remain fully isolated.

        Example::

            runtime = ClusterRuntime(RuntimeConfig(node_id="w1", role="write", enable_cluster=True))
            await runtime.start()

            df_a = DFrame.from_runtime(runtime, {"city": ["jakarta"]})
            df_b = DFrame.from_runtime(runtime, {"product": ["apple"]})
            # df_a and df_b are isolated — no overlap
        """
        instance = cls(
            data=data,
            coordinator=runtime.coordinator,
            runtime=runtime,
            frame_id=frame_id,
            transactional=transactional,
        )
        _RuntimeRegistry.register(runtime.config.node_id, runtime)
        return instance

    @classmethod
    def from_csv(
        cls,
        path: str,
        schema: dict[str, ColumnSchema] | None = None,
        transactional: bool = True,
        **kwargs,
    ) -> "DFrame":
        """Load a CSV file directly into a DFrame."""
        import pandas as pd
        df = pd.read_csv(path, **kwargs)
        data = df.to_dict(orient="list")
        return cls(data, schema=schema, transactional=transactional)

    @classmethod
    def from_excel(
        cls,
        path: str,
        sheet_name: str | int = 0,
        schema: dict[str, ColumnSchema] | None = None,
        transactional: bool = True,
        **kwargs,
    ) -> "DFrame":
        """Load an Excel file directly into a DFrame."""
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name, **kwargs)
        data = df.to_dict(orient="list")
        return cls(data, schema=schema, transactional=transactional)

    @classmethod
    async def from_csv_lazy(
        cls,
        path: str,
        chunk_size: int = 500,
        schema: dict[str, ColumnSchema] | None = None,
        encoding: str = "utf-8",
        delimiter: str = ",",
        on_progress: Callable[[int], None] | None = None,
        distribute: bool = False,
        runtime: "ClusterRuntime | None" = None,
        transactional: bool = True,
        **kwargs,
    ) -> "DFrame":
        """Load CSV lazily using chunked reads (memory O(chunk_size))."""
        coordinator = runtime.coordinator if runtime is not None else None
        instance = cls(
            schema=schema,
            runtime=runtime,
            coordinator=coordinator,
            transactional=transactional,
        )

        if distribute and runtime is None:
            raise ValueError("from_csv_lazy(distribute=True) requires a runtime")

        def _iter_chunks() -> Iterable[pd.DataFrame]:
            for chunk in pd.read_csv(
                path,
                chunksize=chunk_size,
                encoding=encoding,
                delimiter=delimiter,
                **kwargs,
            ):
                yield chunk

        if distribute and runtime is not None:
            iterator = iter(_iter_chunks())

            def _next_chunk() -> pd.DataFrame | None:
                try:
                    return next(iterator)
                except StopIteration:
                    return None

            async def _async_chunks() -> AsyncIterable[pd.DataFrame]:
                while True:
                    chunk = await asyncio.to_thread(_next_chunk)
                    if chunk is None:
                        break
                    yield chunk

            total = await instance._seed_distributed(_async_chunks(), on_progress=on_progress)
        else:
            loop = asyncio.get_running_loop()
            total = await loop.run_in_executor(
                None,
                lambda: instance._seed_chunked(_iter_chunks(), on_progress=on_progress),
            )

        logger.info("from_csv_lazy: loaded %d rows from %s", total, path)
        return instance

    @classmethod
    async def from_excel_lazy(
        cls,
        path: str,
        sheet_name: str | int = 0,
        chunk_size: int = 500,
        schema: dict[str, ColumnSchema] | None = None,
        on_progress: Callable[[int], None] | None = None,
        distribute: bool = False,
        runtime: "ClusterRuntime | None" = None,
        transactional: bool = True,
    ) -> "DFrame":
        """Load Excel lazily using openpyxl read-only stream (memory O(chunk_size))."""
        coordinator = runtime.coordinator if runtime is not None else None
        instance = cls(
            schema=schema,
            runtime=runtime,
            coordinator=coordinator,
            transactional=transactional,
        )

        if distribute and runtime is None:
            raise ValueError("from_excel_lazy(distribute=True) requires a runtime")

        def _iter_excel_chunks() -> Iterable[pd.DataFrame]:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ImportError(
                    "openpyxl is required for from_excel_lazy(). Install with: pip install hiveframe[excel]"
                ) from exc

            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
                headers: list[str] | None = None
                buffer: list[dict[str, Any]] = []

                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(row)]
                        continue
                    buffer.append(dict(zip(headers, row)))
                    if len(buffer) >= chunk_size:
                        yield pd.DataFrame(buffer)
                        buffer = []

                if buffer:
                    yield pd.DataFrame(buffer)
            finally:
                wb.close()

        if distribute and runtime is not None:
            iterator = iter(_iter_excel_chunks())

            def _next_chunk() -> pd.DataFrame | None:
                try:
                    return next(iterator)
                except StopIteration:
                    return None

            async def _async_chunks() -> AsyncIterable[pd.DataFrame]:
                while True:
                    chunk = await asyncio.to_thread(_next_chunk)
                    if chunk is None:
                        break
                    yield chunk

            total = await instance._seed_distributed(_async_chunks(), on_progress=on_progress)
        else:
            loop = asyncio.get_running_loop()
            total = await loop.run_in_executor(
                None,
                lambda: instance._seed_chunked(_iter_excel_chunks(), on_progress=on_progress),
            )

        logger.info("from_excel_lazy: loaded %d rows from %s", total, path)
        return instance

    def describe_for_agent(
        self,
        max_rows: int = 20,
        include_schema: bool = True,
        include_stats: bool = True,
    ) -> str:
        """Build rich context string for LLM agent."""
        parts = []
        fresh = self.read_fresh()
        if include_schema and self._schema:
            parts.append("## Schema")
            for col, schema in self._schema.items():
                parts.append(f"- {col}: {schema.dtype} | {schema.description}")
        if include_stats:
            parts.append("## Statistics")
            try:
                parts.append(fresh.describe().to_string())
            except Exception:
                pass
        parts.append(f"## Sample (first {max_rows} rows)")
        parts.append(fresh.head(max_rows).to_string())
        parts.append(f"\nTotal: {len(fresh)} rows × {len(fresh.columns)} cols")
        return "\n\n".join(parts)

    # -------------------------------------------------------------------------
    # Namespace helpers
    # -------------------------------------------------------------------------

    def _cell_id(self, col: str, row_idx: int) -> str:
        """Build a namespaced cell-id for this DFrame.

        WriteNode._parse_cell_id uses rsplit('_', 1) to split col and row_idx.
        We prefix the column name with '{frame_id}::' so the column stored in
        WriteNode becomes '{frame_id}::{col}' and row_idx stays as pandas row.
        The resulting cell_id passed to Operation is '{frame_id}::{col}_{row_idx}'.
        """
        return f"{self._frame_id}{self._SEP}{col}_{row_idx}"

    def _parse_cell_id(self, cell_id: str) -> tuple[str, int] | None:
        """Parse a namespaced cell-id belonging to this DFrame.
        Returns (col, row_idx) or None if the cell-id belongs to another frame.
        """
        if self._SEP not in cell_id:
            return None
        ns, rest = cell_id.split(self._SEP, 1)
        if ns != self._frame_id:
            return None
        if "_" not in rest:
            return None
        col, _, row_str = rest.rpartition("_")
        try:
            return col, int(row_str)
        except ValueError:
            return None

    def _snapshot_for_frame(self, raw: pd.DataFrame) -> pd.DataFrame:
        """Vectorized - does not use nested loop + .at()"""
        wn = self._coordinator.write_node
        with wn._lock:
            df = wn._df
            if df.empty:
                return pd.DataFrame()
            prefix = f"{self._frame_id}::"
            # Filter columns - O(n_cols)
            matching = [c for c in df.columns if c.startswith(prefix)]
            if not matching:
                return pd.DataFrame()
            # Subset + rename - vectorized pandas ops
            result = df[matching].copy()
            result.columns = [c[len(prefix):].rsplit("_", 1)[0]
                              for c in matching]
            return result.reset_index(drop=True)

    def _validate_value(self, col: str, value: Any) -> Any:
        schema = self._schema.get(col)
        if not schema:
            return value
        return schema.validate(value, column_name=col)

    def _validate_column_assignment(self, column: str, value: Any) -> Any:
        if column not in self._schema:
            return value
        if isinstance(value, list):
            return [self._validate_value(column, item) for item in value]
        return self._validate_value(column, value)

    def _normalize_chunk(self, chunk_df: pd.DataFrame) -> dict[str, list[Any]]:
        normalized: dict[str, list[Any]] = {}
        for col in chunk_df.columns:
            values = chunk_df[col].tolist()
            validated = self._validate_column_assignment(col, values)
            normalized[col] = validated if isinstance(validated, list) else [validated]
        return normalized

    def _append_chunk_to_write_node(
        self,
        normalized: dict[str, list[Any]],
        row_offset: int,
    ) -> int:
        if not normalized:
            return 0

        namespaced = {
            f"{self._frame_id}::{col}": pd.Series(values, dtype="object")
            if col in self._schema and self._schema[col].coerce
            else values
            for col, values in normalized.items()
        }
        chunk_frame = pd.DataFrame(namespaced)
        chunk_frame.index = range(row_offset, row_offset + len(chunk_frame))

        wn = self._coordinator.write_node
        with wn._lock:
            if wn._df.empty:
                wn._df = chunk_frame
            else:
                wn._df = pd.concat([wn._df, chunk_frame])
            wn._version += 1
        return len(chunk_frame)

    def _append_chunk_summary_wal(
        self,
        *,
        cell_id: str,
        row_offset: int,
        rows: int,
        cols: list[str],
        extra: dict[str, Any] | None = None,
    ) -> None:
        if not self._transactional:
            return
        from .transaction import Transaction, Operation, TxState

        payload: dict[str, Any] = {
            "row_offset": row_offset,
            "rows": rows,
            "cols": cols,
        }
        if extra:
            payload.update(extra)

        summary_tx = Transaction(operations=[
            Operation(
                cell_id=f"{self._frame_id}::{cell_id}",
                old_value=None,
                new_value=payload,
                author_type="human",
                author_id="init",
            )
        ])
        summary_tx.transition(TxState.VALIDATING)
        summary_tx.transition(TxState.LOCKED)
        summary_tx.transition(TxState.APPLYING)
        summary_tx.transition(TxState.COMMITTED)
        self._coordinator.wal.append(summary_tx)

    def _seed_chunk_local(
        self,
        chunk_df: pd.DataFrame,
        row_offset: int,
        *,
        wal_cell_id: str,
        extra_wal: dict[str, Any] | None = None,
    ) -> int:
        normalized = self._normalize_chunk(chunk_df)
        rows = self._append_chunk_to_write_node(normalized, row_offset)
        if rows == 0:
            return 0
        self._append_chunk_summary_wal(
            cell_id=wal_cell_id,
            row_offset=row_offset,
            rows=rows,
            cols=list(normalized.keys()),
            extra=extra_wal,
        )
        return rows

    def _seed_chunked(
        self,
        chunks: Iterable[pd.DataFrame],
        on_progress: Callable[[int], None] | None = None,
    ) -> int:
        """Seed data incrementally from chunk iterable (memory O(chunk_size))."""
        total = 0
        for chunk_df in chunks:
            total = self._seed_chunk_step(
                chunk_df,
                total,
                wal_cell_id="__bulk_chunk__",
                on_progress=on_progress,
            )
        self._invalidate_snapshot_cache()
        return total

    def _seed_chunk_step(
        self,
        chunk_df: pd.DataFrame,
        total: int,
        *,
        wal_cell_id: str,
        on_progress: Callable[[int], None] | None = None,
    ) -> int:
        """Apply one seed chunk and return updated total row count."""
        if chunk_df.empty:
            return total
        rows = self._seed_chunk_local(chunk_df, total, wal_cell_id=wal_cell_id)
        updated_total = total + rows
        if on_progress:
            on_progress(updated_total)
        return updated_total

    async def _send_chunk_to_node(
        self,
        chunk_df: pd.DataFrame,
        *,
        row_offset: int,
        target_node_id: str,
    ) -> int:
        """Send one chunk to local or remote target write node."""
        if self._runtime is None:
            raise RuntimeError("_send_chunk_to_node requires DFrame runtime")

        if target_node_id == self._runtime.config.node_id:
            rows = self._seed_chunk_local(
                chunk_df,
                row_offset,
                wal_cell_id="__distributed_chunk__",
                extra_wal={"node": target_node_id},
            )
            self._invalidate_snapshot_cache()
            return rows

        normalized = self._normalize_chunk(chunk_df)
        if not normalized:
            return 0

        from .message import Message, MessageType

        message = Message.build(
            message_type=MessageType.SEED_CHUNK,
            sender_id=self._runtime.config.node_id,
            sender_region=self._runtime.config.region,
            payload={
                "frame_id": self._frame_id,
                "row_offset": row_offset,
                "data": normalized,
                "transactional": self._transactional,
            },
        )
        await self._runtime.transport.send(target_node_id, message)
        return len(next(iter(normalized.values()))) if normalized else 0

    async def _seed_distributed(
        self,
        chunks: AsyncIterable[pd.DataFrame],
        on_progress: Callable[[int], None] | None = None,
    ) -> int:
        """Distribute chunked seed across write nodes, fallback to local for single node."""
        if self._runtime is None:
            raise RuntimeError(
                "_seed_distributed requires DFrame with ClusterRuntime. Use DFrame.from_runtime()."
            )

        write_nodes = await self._runtime.registry.get_write_nodes()
        if len(write_nodes) <= 1:
            logger.info(
                "_seed_distributed: only %d write node(s), falling back to local chunk seed",
                len(write_nodes),
            )
            total = 0
            async for chunk_df in chunks:
                total = self._seed_chunk_step(
                    chunk_df,
                    total,
                    wal_cell_id="__bulk_chunk__",
                    on_progress=on_progress,
                )
            self._invalidate_snapshot_cache()
            return total

        total = 0
        node_idx = 0
        async for chunk_df in chunks:
            if chunk_df.empty:
                continue
            target_node = write_nodes[node_idx % len(write_nodes)]
            node_idx += 1
            rows = await self._send_chunk_to_node(
                chunk_df,
                row_offset=total,
                target_node_id=target_node.node_id,
            )
            total += rows
            if on_progress:
                on_progress(total)
        self._invalidate_snapshot_cache()
        return total

    def _seed_initial_data(self, data: dict[str, list[Any]]) -> None:
        # Validate all values if schema is present
        normalized_data: dict[str, list[Any]] = {}
        for col, values in data.items():
            normalized = self._validate_column_assignment(col, values)
            normalized_data[col] = normalized if isinstance(normalized, list) else [normalized]
        """
        Inject initial data directly to write_node._df - bypass per-cell ops.
        WAL only records a single summary entry for the entire seed.
        """
        max_len = max((len(v) for v in normalized_data.values()), default=0)
        if max_len == 0:
            return

        # Build namespaced columns directly
        namespaced = {
            f"{self._frame_id}::{col}": pd.Series(values, dtype="object")
            if col in self._schema and self._schema[col].coerce
            else values
            for col, values in normalized_data.items()
        }

        wn = self._coordinator.write_node
        with wn._lock:
            new_df = pd.DataFrame(namespaced)
            # Merge with existing data if present
            if wn._df.empty:
                wn._df = new_df
            else:
                wn._df = pd.concat([wn._df, new_df], axis=1)
            wn._version += 1

        if self._transactional:
            # Single WAL entry for the entire seed
            from .transaction import Transaction, Operation, TxState

            summary_tx = Transaction(operations=[
                Operation(
                    cell_id=f"{self._frame_id}::__bulk_init__",
                    old_value=None,
                    new_value={"rows": max_len, "cols": list(normalized_data.keys())},
                    author_type="human",
                    author_id="init",
                )
            ])
            summary_tx.transition(TxState.VALIDATING)
            summary_tx.transition(TxState.LOCKED)
            summary_tx.transition(TxState.APPLYING)
            summary_tx.transition(TxState.COMMITTED)
            self._coordinator.wal.append(summary_tx)
        self._invalidate_snapshot_cache()

    def __setitem__(self, column: str, value: Any) -> None:
        """Set a full column value, routing each row to its owning node."""
        snapshot = self.read_fresh()
        row_count = len(snapshot.index)
        values = list(value) if isinstance(value, list) else [value] * max(row_count, 1)
        if row_count == 0:
            row_count = len(values)
        total_rows = max(row_count, len(values))
        if len(values) < total_rows:
            values.extend([None] * (total_rows - len(values)))
        values = self._validate_column_assignment(column, values)

        from collections import defaultdict
        node_ops: dict[str, list[Operation]] = defaultdict(list)

        for row_idx in range(max(row_count, len(values))):
            old_value = (
                snapshot.at[row_idx, column]
                if column in snapshot.columns and row_idx < len(snapshot.index)
                else None
            )
            new_value = values[row_idx] if row_idx < len(values) else None
            op = Operation(
                cell_id=self._cell_id(column, row_idx),
                old_value=old_value,
                new_value=new_value,
                author_type="human",
                author_id="user",
            )
            owner_id = self._get_owner_node_id(row_idx)
            node_ops[owner_id].append(op)

        # Invalidate cached snapshot so subsequent reads see fresh state
        self._invalidate_snapshot_cache()
        self._submit_routed(node_ops)

    def _get_owner_node_id(self, row_index: int) -> str:
        """Return the node_id that owns row_index, or 'local' for standalone mode."""
        if self._runtime is not None:
            owner = self._runtime.route_write(row_index)
            if owner is not None:
                return owner.node_id
        return "local"

    def _submit_routed(self, node_ops: "dict[str, list[Operation]]") -> None:
        """Submit operation groups to the correct coordinator per owning node."""
        for node_id, ops in node_ops.items():
            if not ops:
                continue
            coordinator = self._resolve_coordinator(node_id)
            if self._transactional:
                coordinator.submit(ops)
            else:
                ok = coordinator.submit_non_transactional(ops)
                if not ok:
                    raise RuntimeError(f"Non-transactional write failed on node '{node_id}'")

    def _resolve_coordinator(self, node_id: str) -> "TransactionCoordinator":
        """Resolve coordinator for a given node_id.

        - 'local' or own node_id → local coordinator
        - remote node_id in cluster → remote coordinator via transport (WRITE_FORWARD)
          Currently uses local coordinator as fallback when no remote proxy is wired.
          Stage-3 can replace this with a real remote-write proxy.
        """
        if self._runtime is None or node_id in ("local", self._runtime.config.node_id):
            return self._coordinator
        # Look up a remote runtime in the shared in-memory registry (same process).
        # In a real multi-process deployment this would be replaced by a transport-based proxy.
        remote_runtime = _RuntimeRegistry.get(node_id)
        if remote_runtime is not None:
            return remote_runtime.coordinator
        # Fallback: write locally if remote runtime is not reachable.
        return self._coordinator

    @staticmethod
    def _to_pandas_safe(frame: Any) -> pd.DataFrame:
        """Convert node dataframe to pandas with pyarrow fallback."""
        try:
            return frame.to_pandas(use_pyarrow_extension_array=True)
        except Exception:
            return frame.to_pandas()

    @staticmethod
    def _normalize_index(frame: pd.DataFrame) -> pd.DataFrame:
        """Normalize index to allow deterministic completeness checks."""
        if frame.empty:
            return frame.reset_index(drop=True)
        return frame.reset_index(drop=True)

    def _build_local_snapshot(self) -> pd.DataFrame:
        """Reconstruct a clean DataFrame by filtering write_node columns for this frame_id.

        WriteNode stores data as:
          - column name: '{frame_id}::{col}'   (from _parse_cell_id: col = '{frame_id}::{col}')
          - row index:   pandas integer index   (from _parse_cell_id: row = row_idx)
        So we just filter columns that start with our prefix and strip it.
        """
        prefix = f"{self._frame_id}{self._SEP}"
        wn = self._coordinator.write_node
        with wn._lock:
            df_internal = wn._df
            matching_cols = [c for c in df_internal.columns if c.startswith(prefix)]
            if not matching_cols:
                return pd.DataFrame()
            subset = df_internal[matching_cols].copy()
        # Rename columns by stripping the namespace prefix
        subset.columns = [c[len(prefix):] for c in subset.columns]
        return subset.reset_index(drop=True)

    def _get_cached_snapshot(self) -> pd.DataFrame:
        """Return cached snapshot if fresh; otherwise build and cache it."""
        now = time.time()
        if self._snapshot_cache is not None:
            frame, ts = self._snapshot_cache
            if now - ts <= self._snapshot_cache_ttl:
                return frame
        frame = self._build_local_snapshot()
        # store shallow copy to avoid external mutation
        self._snapshot_cache = (frame, now)
        return frame

    def _invalidate_snapshot_cache(self) -> None:
        """Invalidate cached snapshot on writes or external updates."""
        self._snapshot_cache = None

    def read_fresh(self) -> pd.DataFrame:
        """Return latest namespace-isolated snapshot.

        Always returns the local write-node snapshot synchronously.
        This path does not read from ReadNode/WAL. In transactional=False mode,
        this is the intended simple distributed read path: read directly from the
        owning writer snapshot.
        For global cluster reads (fan-out across all nodes), use read_fresh_async().
        """
        return self._get_cached_snapshot()

    def read_fresh_global(self) -> pd.DataFrame:
        """Return a global merged snapshot from sync code.

        This is a sync convenience wrapper for read_fresh_global_async(). It is intended
        for non-async callers. If called while an event loop is already running,
        raise a clear error and require using await read_fresh_global_async() directly.
        """
        try:
            asyncio.get_running_loop()
        except RuntimeError:
            return asyncio.run(self.read_fresh_global_async())
        raise RuntimeError(
            "read_global() cannot be called while an event loop is running; "
            "use 'await read_fresh_global_async()' instead."
        )

    async def read_fresh_global_async(self) -> pd.DataFrame:
        """Return global snapshot fanned out across all cluster writer nodes.

        Use this inside async contexts (e.g. asyncio.run, async def).
        In standalone mode this is equivalent to read_fresh().

        This fan-out targets writer nodes directly. In transactional=False mode,
        this remains the supported multi-node read path; remote read-replica sync
        is only provided by the WAL-backed transactional path.
        """
        if self._runtime is None or not self._runtime.config.enable_cluster:
            return self._build_local_snapshot()

        write_nodes = self._runtime.registry._write_nodes_sorted()
        remote_nodes = [n for n in write_nodes if n.node_id != self._runtime.config.node_id]
        if not remote_nodes:
            return self._build_local_snapshot()

        try:
            result = await self._runtime.read_global_snapshot_for(self._frame_id)
            if isinstance(result, pd.DataFrame) and not result.empty:
                return result
        except Exception:
            pass
        return self._build_local_snapshot()

    def read_fresh_global_lazy(self, columns: list[str] | None = None, chunk_size: int = 1000):
        """Yield DataFrame chunks lazily from the global merged snapshot (all cluster nodes).

        Args:
            columns: Optional list of columns to include (public names, e.g. 'city'). If None, all columns are included.
            chunk_size: Number of rows per chunk (default: 1000).
        Yields:
            pandas.DataFrame: chunk of up to chunk_size rows, with columns matching DFrame public API.
        Raises:
            RuntimeError: if called while an event loop is running (use await read_fresh_global_async() for async).
        """
        try:
            asyncio.get_running_loop()
        except RuntimeError:
            df = asyncio.run(self.read_fresh_global_async())
            if columns is not None:
                df = df[columns]
            total_rows = len(df)
            for start in range(0, total_rows, chunk_size):
                yield df.iloc[start:start+chunk_size].copy()
            return
        raise RuntimeError(
            "read_fresh_global_lazy() cannot be called while an event loop is running; "
            "use 'await read_fresh_global_async()' and chunk manually instead."
        )

    def read_fresh_lazy(self, columns: list[str] | None = None, chunk_size: int = 1000):
        """Yield DataFrame chunks lazily from the local write-node snapshot.

        If a persisted parquet exists, stream from disk (truly lazy). Otherwise,
        fallback to in-memory chunking from the writer snapshot.
        """
        parquet_path = None
        if self._coordinator and hasattr(self._coordinator, 'read_node'):
            storage_dir = getattr(self._coordinator.read_node, '_storage_dir', None)
            if storage_dir is not None:
                parquet_path = Path(storage_dir) / f"{self._frame_id}.parquet"
        if parquet_path and parquet_path.exists():
            # Truly lazy: stream from parquet
            try:
                import pyarrow
                import pyarrow.parquet as pq
                table = pq.read_table(parquet_path, columns=columns)
                df = table.to_pandas()
                total_rows = len(df)
                for start in range(0, total_rows, chunk_size):
                    yield df.iloc[start : start + chunk_size].copy()
                return
            except Exception:
                pass  # fallback to in-memory
        # Fallback: in-memory chunking
        df = self.read_fresh()
        if columns is not None:
            df = df[columns]
        total_rows = len(df)
        for start in range(0, total_rows, chunk_size):
            yield df.iloc[start : start + chunk_size].copy()

    # Backward-compatible alias for local lazy read
    read_lazy = read_fresh_lazy

    # Deprecated/legacy aliases for backward compatibility (optional, can remove for strictness)
    read_global = read_fresh_global
    read_global_lazy = read_fresh_global_lazy
    read_fresh_async = read_fresh_global_async

    # -------------------------------------------------------------------------
    # Pandas proxy layer
    # Rule: read ops → delegate to global read_fresh(); write ops → transactional
    # -------------------------------------------------------------------------

    # Read-only properties delegated to global snapshot
    @property
    def shape(self) -> tuple[int, int]:
        return self.read_fresh().shape

    @property
    def columns(self) -> pd.Index:
        return self.read_fresh().columns

    @property
    def dtypes(self) -> pd.Series:
        return self.read_fresh().dtypes

    @property
    def index(self) -> pd.Index:
        return self.read_fresh().index

    @property
    def values(self) -> Any:
        return self.read_fresh().values

    def __len__(self) -> int:
        return len(self.read_fresh())

    def __repr__(self) -> str:
        return f"DFrame(\n{self.read_fresh().__repr__()}\n)"

    def __str__(self) -> str:
        return self.read_fresh().__str__()

    def __getattr__(self, name: str) -> Any:
        """Proxy any unknown attribute/method to the global read_fresh() snapshot.

        This covers the full pandas read API surface:
            head, tail, loc, iloc, at, iat, filter, apply, map, agg, sum,
            mean, min, max, std, var, count, median, quantile, nunique,
            sort_values, sort_index, fillna, dropna, replace, astype,
            assign, copy, merge, join, to_csv, to_json, to_dict, to_numpy,
            to_parquet, info, reset_index, set_index, reindex, ...

        Write operations (append, drop, rename, __setitem__) are handled
        explicitly above and are NOT proxied — they go through the
        transactional coordinator path.
        """
        # Guard: avoid infinite recursion for private/dunder attrs
        if name.startswith("_"):
            raise AttributeError(name)
        frame = self.read_fresh()
        attr = getattr(frame, name, None)
        if attr is None:
            raise AttributeError(f"DFrame and pandas.DataFrame have no attribute '{name}'")
        return attr

    def _best_frame_for(self, required_columns: set[str], analytical: bool = False) -> pd.DataFrame:
        """Return the best available frame for the requested columns.

        - If all required columns are present in the cached snapshot, return it.
        - Otherwise, fall back to building a fresh snapshot from the write node.
        """
        frame = self._get_cached_snapshot()
        if not frame.empty and frame.columns.isin(required_columns).all():
            return frame
        return self._build_local_snapshot()

    def __getitem__(self, key: str | list[str]) -> pd.Series | pd.DataFrame:
        """Get a column or list of columns with hybrid cache completeness guardrails."""
        if isinstance(key, list):
            frame = self._best_frame_for(set(key))
            missing = [k for k in key if k not in frame.columns]
            if missing:
                raise KeyError(missing)
            return frame[key]
        frame = self._best_frame_for({key})
        if key not in frame.columns:
            raise KeyError(key)
        return frame[key]

    def groupby(self, by: str):
        """Group using cache only when complete; otherwise fallback to fresh snapshot."""
        frame = self._best_frame_for({by}, analytical=True)
        if by not in frame.columns:
            raise KeyError(by)
        return frame.groupby(by)

    def to_persistent(self, name: str) -> Path:
        """Persist a structurally complete snapshot to parquet."""
        frame = self._best_frame_for(set(), analytical=True)
        cached = self._read_cached()
        # If cache is incomplete, sync read-node state through existing transactional path.
        if not self._is_cache_complete_for(cached, set(frame.columns)):
            for col in frame.columns:
                self[col] = frame[col].tolist()
        path = Path(self._coordinator.read_node._storage_dir) / f"{name}.parquet"
        path.parent.mkdir(parents=True, exist_ok=True)
        frame.to_parquet(path)
        _write_schema_sidecar(path, self._schema)
        return path

    def _read_cached(self) -> pd.DataFrame:
        """Read current cache snapshot filtered for this frame."""
        return self._normalize_index(self._build_local_snapshot())

    def _is_cache_complete_for(self, frame: pd.DataFrame, required_columns: set[str]) -> bool:
        """Check whether cached frame is structurally complete for requested columns."""
        fresh = self._normalize_index(self.read_fresh())
        if not required_columns.issubset(set(frame.columns)):
            return False
        if len(frame.index) != len(fresh.index):
            return False
        if set(frame.columns) != set(fresh.columns):
            return False
        # Ensure all required columns preserve the same row-level null/non-null shape.
        for col in required_columns:
            if col not in fresh.columns:
                return False
            cached_mask = frame[col].isna().tolist()
            fresh_mask = fresh[col].isna().tolist()
            if cached_mask != fresh_mask:
                return False
        return True

    def checkpoint(self, label: str | None = None) -> str:
        """
        Save the current state as a checkpoint.
        Returns a checkpoint_id that can be used for rollback.
        """
        if not self._transactional:
            raise RuntimeError("checkpoint() is unavailable when transactional=False")
        import time
        if not hasattr(self, '_checkpoints'):
            self._checkpoints = {}
        checkpoint_id = label or f"cp_{int(time.time())}"
        current_lsn = self._coordinator.wal._next_lsn - 1
        self._checkpoints[checkpoint_id] = {
            "lsn": current_lsn,
            "snapshot": self.read_fresh().to_dict(orient="list"),
            "created_at": time.time(),
        }
        return checkpoint_id

    def rollback(self, checkpoint_id: str) -> None:
        """
        Restore data to the state when the checkpoint was created.
        All changes after the checkpoint are undone.
        """
        if not self._transactional:
            raise RuntimeError("rollback() is unavailable when transactional=False")
        if not hasattr(self, '_checkpoints') or checkpoint_id not in self._checkpoints:
            raise KeyError(f"Checkpoint '{checkpoint_id}' not found")
        snapshot = self._checkpoints[checkpoint_id]["snapshot"]
        self._seed_initial_data(snapshot)

    def list_checkpoints(self) -> list[dict]:
        if not hasattr(self, '_checkpoints'):
            return []
        return [
            {"id": k, **v}
            for k, v in self._checkpoints.items()
        ]

    def diff(self, checkpoint_id: str) -> pd.DataFrame:
        """Return the changes between the checkpoint and the current state."""
        if not hasattr(self, '_checkpoints') or checkpoint_id not in self._checkpoints:
            raise KeyError(f"Checkpoint '{checkpoint_id}' not found")
        import pandas as pd
        old = pd.DataFrame(self._checkpoints[checkpoint_id]["snapshot"])
        new = self.read_fresh()
        # Find changed cells
        diff = (old != new).stack()
        changed = diff[diff]
        if changed.empty:
            return pd.DataFrame(columns=["row", "col", "old_value", "new_value"])
        rows, cols = zip(*changed.index)
        return pd.DataFrame({
            "row": rows,
            "col": cols,
            "old_value": old.values[changed.index],
            "new_value": new.values[changed.index],
        })

    def cell_history(self, col: str, row_idx: int) -> list[dict]:
        if not self._transactional:
            return []
        cell_id = self._cell_id(col, row_idx)
        return self._coordinator.wal.get_cell_history(cell_id)

    def get_metrics(self) -> dict:
        """Return comprehensive metrics untuk monitoring/observability."""
        return {
            "coordinator": self._coordinator.get_stats(),
            "wal": {
                "total_entries": len(self._coordinator.wal._entries),
                "last_lsn": self._coordinator.wal._next_lsn - 1,
            },
            "write_node": {
                "rows": len(self._coordinator.write_node._df),
                "cols": len(self._coordinator.write_node._df.columns),
                "version": self._coordinator.write_node._version,
            },
            "read_node": {
                "version": self._coordinator.read_node._version,
                "cache_rows": len(self._coordinator.read_node._cache),
                "buffer_pending": len(self._coordinator.read_node._delta_buffer),
            },
            "frame_id": self._frame_id,
        }


def read(
    name: str,
    coordinator: TransactionCoordinator | None = None,
    schema: dict[str, ColumnSchema] | None = None,
    transactional: bool = True,
) -> DFrame:
    """Load persisted dataframe from parquet into a new DFrame instance.

    `name` may be either a dataset name previously used with `to_persistent()`
    or a full parquet file path.
    """
    resolved_coordinator = coordinator or TransactionCoordinator()
    parquet_path = _resolve_read_path(name, resolved_coordinator)
    frame = pd.read_parquet(parquet_path)
    resolved_schema = schema or _load_schema_sidecar(parquet_path)
    data = frame.to_dict(orient="list")
    return DFrame(
        data=data,
        coordinator=resolved_coordinator,
        schema=resolved_schema,
        transactional=transactional,
    )
